"""Generate LED_MUX_CONTROLLER_testplan.xlsx from TESTPLAN.md content."""
import argparse
import re
import sys
import openpyxl
from openpyxl.styles import Alignment

OUT = r"c:\Users\chani\AI for me\dv_project\LED_MUX_CONTROLLER_testplan.xlsx"
TEMPLATE = r"c:\Users\chani\AI for me\dv_project\LED_MUX_CONTROLLER_stu\Template-TestPlan.xlsx"
PLAN_NAME = "led_mux_controller_testplan"

# Test rank per TESTPLAN.md §1 (P0 essential, P1 good-to-have, P2 stretch)
TEST_RANK = {
    "smoke_test": "P0",
    "apb_reset_defaults_test": "P0",
    "apb_led_enable_write_read_test": "P0",
    "apb_scratchpad_wr_rd_test": "P0",
    "apb_invalid_addr_test": "P0",
    "apb_pready_no_wait_test": "P0",
    "led_reset_values_test": "P0",
    "led_decimal_42_test": "P0",
    "led_overflow_modulo_test": "P0",
    "led_disable_blocks_update_test": "P0",
    "led_all_digits_0_to_9_test": "P0",
    "apb_default_enable_led_path_test": "P1",
    "apb_read_during_processing_test": "P1",
    "led_max_displayable_test": "P1",
    "led_sel_onehot_scan_test": "P1",
    "led_hold_time_min_test": "P1",
    "led_latency_window_test": "P1",
    "led_reenable_after_disable_test": "P1",
    "full_display_flow_test": "P1",
    "random_regression_test": "P1",
    "apb_done_read_only_test": "P2",
    "apb_done_poll_timeout_test": "P2",
    "apb_scratchpad_all_ones_test": "P2",
    "apb_scratchpad_walking_one_test": "P2",
    "apb_enable_toggle_stress_test": "P2",
    "led_single_digit_zero_test": "P2",
    "led_single_digit_one_test": "P2",
    "led_seg_active_low_test": "P2",
    "led_overflow_max_test": "P2",
    "led_overflow_boundary_test": "P2",
    "led_back_to_back_error_test": "P2",
    "led_reset_during_display_test": "P2",
    "led_done_clear_after_reset_test": "P2",
    "led_hold_below_min_negative_test": "P2",
    "scratch_then_display_test": "P2",
    "virtual_seq_stress_test": "P2",
    "enable_off_overflow_test": "P2",
    "poll_until_done_stress_test": "P2",
}

TIER_LEVEL = {"p0": 0, "p1": 1, "p2": 2, "all": 3}
RANK_LEVEL = {"P0": 0, "P1": 1, "P2": 2}


def filter_rows_by_tier(rows, tier: str):
    tier = tier.lower()
    if tier not in TIER_LEVEL:
        raise ValueError(f"Unknown tier {tier!r}; use p0, p1, p2, or all")
    if tier == "all":
        return rows
    max_level = TIER_LEVEL[tier]
    return [
        r for r in rows
        if RANK_LEVEL.get(TEST_RANK.get(r[2], "P2"), 2) <= max_level
    ]

ALL_SVA_PROPERTIES = (
    "assert_sel_out_reset_value, assert_seg_out_reset_value, "
    "assert_sel_out_onehot_active_low, assert_seg_out_bit7_always_one, "
    "assert_apb_setup_phase, assert_apb_access_phase, assert_apb_pready_complete, "
    "assert_apb_pslerr_invalid_addr, cover_sel_out_digit_position, "
    "cover_seg_out_decimal_digit, check_60_80_cycle, check_hold_1002_cycle"
)


def parse_checkers(ch: str):
    """Parse checker tags into SVA, covergroup, and scoreboard lists.

    Monitors observe only; pass/fail checking belongs in SVA or scoreboard.
    """
    sva, cov, scb = [], [], []
    text = ch.replace("**", "")
    if re.search(r"\bMON\b", text):
        raise ValueError(
            "MON checker tags are not allowed — use SVA or scoreboard (SCB-*): "
            f"{ch!r}"
        )
    if "all §0.3" in text.lower() or ALL_SVA_PROPERTIES in text:
        sva = [ALL_SVA_PROPERTIES]
    else:
        for m in re.findall(
            r"(?:assert|cover|check)_[a-z0-9_]+", text
        ):
            if m not in sva:
                sva.append(m)
    for m in re.findall(r"cg_\w+", text):
        cov.append(m)
    if "COV all" in text or "all covergroups" in text.lower():
        cov = ["cg_error_q, cg_digits, cg_enable, cg_done, cg_overflow"]
    for m in re.findall(r"SCB-[\d.]+", text):
        scb.append(m)
    if "SCB-1..9" in text:
        scb = ["SCB-1..9"]

    def dedupe(lst):
        out = []
        for x in lst:
            if x and x not in out:
                out.append(x)
        return out

    return dedupe(sva), dedupe(cov), dedupe(scb)


def fmt_desc(flow: str, constraints: str, checkers: str) -> str:
    """Format $description without spurious '.' lines.

    Flow strings use numbered steps like '1. Reset DUT. 2. Read 0x4000.'
    Split only on whitespace before the next step number (N.), not on every
    period — otherwise '1.' is split from its text and produces lines like '1) .'.
    """
    lines = []
    flow = flow.strip()
    if re.search(r"\d+\.\s", flow):
        parts = re.split(r"\s+(?=\d+\.\s)", flow)
    else:
        parts = [p.strip() for p in re.split(r"(?<=\.)\s+", flow) if p.strip()]

    step_num = 0
    for part in parts:
        text = re.sub(r"^\d+\.\s*", "", part.strip()).rstrip(".")
        if not text or text == ".":
            continue
        step_num += 1
        lines.append(f"{step_num}) {text}.")
    if constraints and constraints.strip() not in ("—", "-"):
        c = constraints.rstrip(".")
        lines.append(f"Constraints: {c}.")
    if checkers:
        ch = checkers.replace("**", "").rstrip(".")
        lines.append(f"Checkers: {ch}.")
    return "\n".join(lines)


def validate_description(desc: str, test_name: str) -> None:
    """Fail fast if $description contains orphan '.' lines."""
    for line in desc.splitlines():
        body = line.strip()
        if re.fullmatch(r"\d+\)\s*\.?", body):
            raise ValueError(
                f"{test_name}: $description has orphan '.' line: {body!r}. "
                "Check fmt_desc step splitting."
            )


# (feature, sub_feature, test, flow, constraints, checkers, code_cov, type, priority, goal)
ROWS = [
    ("General and basic sequence", "Reset values on LED interface", "led_reset_values_test",
     "1. Assert rst_n=0. 2. Check outputs. 3. Deassert reset. 4. Check outputs",
     "Sample on posedge clk after reset (FR 3.3, AC-R1, AC-R2)",
     "assert_sel_out_reset_value, assert_seg_out_reset_value", "line, toggle", "BASIC", "P0", "AC-R1, AC-R2"),

    ("APB Controller", "Reset register defaults", "apb_reset_defaults_test",
     "1. Reset DUT. 2. Read 0x4000 expect LED_enable=1. 3. Read 0x4004 expect Done=0. 4. Read 0x4008 expect 0",
     "No APB traffic during reset assert",
     "SCB-1, SCB-2, SCB-3; assert_sel_out_reset_value, assert_seg_out_reset_value", "line, branch", "BASIC", "P0", "AC-R1, AC-R2"),
    ("", "LED_enable write and read", "apb_led_enable_write_read_test",
     "1. Write 0x4000=0. 2. Read 0x4000. 3. Write 0x4000=1. 4. Read 0x4000",
     "wdata[0] only (FR 3.4, §4.3)",
     "SCB-1; COV cg_enable", "line", "BASIC", "P1", "AC-E1, AC-E2"),
    ("", "Scratch pad write and read", "apb_scratchpad_wr_rd_test",
     "1. Write 0x4008=32'hDEAD_BEEF. 2. Read 0x4008",
     "Full 32-bit wdata (FR 3.7, AC-A1)",
     "SCB-3", "line", "BASIC", "P1", "AC-A1"),
    ("", "Done register read-only", "apb_done_read_only_test",
     "1. Write 0x4004=32'hFFFF_FFFF. 2. Read 0x4004",
     "Done is RO; sequence drives the illegal write (C-7, FR 3.5)",
     "SCB-2", "line, cond", "BASIC", "P1", "AC-D2"),
    ("", "APB pready no-wait-state", "apb_pready_no_wait_test",
     "1. Write 0x4000. 2. Read 0x4004",
     "Standard 2-phase APB (C-6, AC-A3)",
     "assert_apb_setup_phase, assert_apb_access_phase, assert_apb_pready_complete", "line, branch", "BASIC", "P1", "AC-A3"),
    ("", "Default LED_enable display path", "apb_default_enable_led_path_test",
     "1. No write to 0x4000. 2. Drive error_q=7. 3. Poll Done. 4. Scoreboard compare",
     "error_q in [1:99]; default LED_enable=1 (FR 3.4)",
     "SCB-1, SCB-4, SCB-5, SCB-6, SCB-8; assert_sel_out_onehot_active_low, assert_seg_out_bit7_always_one", "line, fsm", "BASIC", "P0", "AC-E1, AC-B1"),

    ("Corner cases", "Invalid APB address", "apb_invalid_addr_test",
     "1. Write 0x5000. 2. Check pslerr. 3. Read 0x0000. 4. Check pslerr",
     "addr NOT IN {0x4000, 0x4004, 0x4008} (AC-A2)",
     "assert_apb_pslerr_invalid_addr", "branch, cond", "CORNER", "P1", "AC-A2"),
    ("", "Done poll timeout without error_q", "apb_done_poll_timeout_test",
     "1. Do not drive error_q. 2. Poll 0x4004 with short timeout",
     "Poll interval >=1 cycle; expect Done==0 (FR 3.5, SCB-8)",
     "SCB-8", "cond", "CORNER", "P2", "AC-D1"),
    ("", "Scratch pad all-ones", "apb_scratchpad_all_ones_test",
     "1. Write 0x4008=32'hFFFF_FFFF. 2. Read back",
     "All bits toggled (FR 3.7)",
     "SCB-3", "toggle", "CORNER", "P2", "AC-A1"),
    ("", "Scratch pad walking-one", "apb_scratchpad_walking_one_test",
     "For i in 0..31: write 1<<i to 0x4008, read back",
     "32 iterations (FR 3.7, AC-A1)",
     "SCB-3", "toggle", "CORNER", "P2", "AC-A1"),
    ("", "LED_enable toggle stress", "apb_enable_toggle_stress_test",
     "1. Toggle LED_enable 5x. 2. Each time drive new error_q. 3. Poll Done when enabled",
     "Distinct error_q per iteration (FR 3.4)",
     "SCB-1, SCB-7; COV cg_enable", "line, cond", "CORNER", "P2", "AC-E2"),
    ("", "Read Done during LED processing", "apb_read_during_processing_test",
     "1. Drive error_q=42. 2. Read 0x4004 before 60 cycles. 3. Read again after poll passes",
     "error_q=42; early read at cycle 10 (C-5, FR 3.5)",
     "SCB-2, SCB-8; check_60_80_cycle", "cond, fsm", "CORNER", "P2", "AC-D1"),

    ("LED MUX", "Single digit zero display", "led_single_digit_zero_test",
     "1. Ensure LED_enable=1. 2. Drive error_q=0. 3. Poll Done. 4. Compare all digit positions",
     "error_q==0 (FR 3.1, AC-B1)",
     "SCB-4, SCB-5, SCB-6; COV cg_digits", "line", "BASIC", "P1", "AC-B1"),
    ("", "Single digit one display", "led_single_digit_one_test",
     "1. Ensure LED_enable=1. 2. Drive error_q=1. 3. Poll Done. 4. Compare segment encoding for digit 1",
     "error_q==1; segments per SPEC §4.4 (AC-B1)",
     "SCB-4, SCB-5, SCB-6; COV cg_digits", "line", "BASIC", "P1", "AC-B1"),
    ("", "Binary-to-BCD display (decimal 42)", "led_decimal_42_test",
     "1. LED_enable=1. 2. Drive error_q=42, hold >=1002 cycles. 3. Poll Done. 4. Scoreboard compare digits 0,0,0,0,4,2",
     "error_q==42; C-4 hold, C-5 latency (AC-B1)",
     "SCB-4, SCB-5, SCB-6, SCB-8; assert_sel_out_onehot_active_low, assert_seg_out_bit7_always_one, check_60_80_cycle; COV cg_digits, cg_error_q", "line, fsm, toggle", "BASIC", "P0", "AC-B1, AC-M1"),
    ("", "Max displayable value (999999)", "led_max_displayable_test",
     "1. Drive error_q=999_999. 2. Poll Done. 3. Compare all six digits show 9",
     "error_q==999_999 (FR 3.1, AC-B1)",
     "SCB-4, SCB-5, SCB-6; assert_sel_out_onehot_active_low, assert_seg_out_bit7_always_one; COV cg_digits", "line", "BASIC", "P1", "AC-B1"),
    ("", "sel_out one-hot multiplex scan", "led_sel_onehot_scan_test",
     "1. Drive error_q=74565. 2. Hold >=1002 cycles. 3. Collect 6+ sel_out samples",
     "Hold 1002 cycles (C-4, AC-M1, AC-M2)",
     "SCB-5; assert_sel_out_onehot_active_low, cover_sel_out_digit_position; COV cg_digits", "fsm, toggle", "BASIC", "P1", "AC-M1, AC-M2, AC-M3"),
    ("", "seg_out active-low encoding", "led_seg_active_low_test",
     "1. Drive error_q=8. 2. Poll Done. 3. Check seg_out[6:0] active-low where lit",
     "error_q==8 (AC-B2, AC-B3)",
     "SCB-6; assert_seg_out_bit7_always_one", "line, toggle", "BASIC", "P1", "AC-B2, AC-B3"),
    ("", "Minimum hold time (1002 cycles)", "led_hold_time_min_test",
     "1. Drive error_q. 2. Hold exactly 1002 cycles. 3. Poll Done",
     "Hold error_q for 1002 cycles; sequence enforces hold duration (C-4, AC-M2)",
     "check_hold_1002_cycle; SCB-4", "line, fsm", "BASIC", "P1", "AC-M2, AC-M3"),
    ("", "Output latency window (60-80 cycles)", "led_latency_window_test",
     "1. Drive error_q=100. 2. Wait 60-80 cycles. 3. Sample outputs",
     "Sequence waits >=60 cycles before sampling (C-5, AC-E1)",
     "check_60_80_cycle; SCB-8", "fsm", "BASIC", "P1", "AC-E1, C-5"),

    ("Corner cases", "Overflow modulo display", "led_overflow_modulo_test",
     "1. Drive error_q=1_000_001. 2. Poll Done. 3. Expect display 000001",
     "Golden = error_q % 1_000_000 (FR 3.6, C-3, AC-B4)",
     "SCB-9; assert_sel_out_onehot_active_low, assert_seg_out_bit7_always_one; COV cg_overflow, cg_digits", "line, branch", "CORNER", "P0", "AC-B4"),
    ("", "Overflow maximum error_q", "led_overflow_max_test",
     "1. Drive error_q=1_048_575. 2. Poll Done. 3. Expect modulo 48_575",
     "error_q==20'hF_FFFF (C-2, AC-B4)",
     "SCB-9; COV cg_overflow", "branch", "CORNER", "P1", "AC-B4"),
    ("", "Overflow boundary values", "led_overflow_boundary_test",
     "Run back-to-back: error_q=999_999, 1_000_000, 1_000_001; poll Done after each",
     "Three explicit boundary values (C-3, AC-B4)",
     "SCB-9; COV cg_overflow", "branch, cond", "CORNER", "P1", "AC-B4"),
    ("", "LED_enable disables propagation", "led_disable_blocks_update_test",
     "1. Write LED_enable=0. 2. Drive error_q=55. 3. Confirm seg_out does not update",
     "LED_enable==0 (FR 3.4, AC-E2)",
     "SCB-7; COV cg_enable", "cond", "CORNER", "P1", "AC-E2"),
    ("", "Re-enable after disable", "led_reenable_after_disable_test",
     "1. Disable. 2. Drive error_q (no effect). 3. Enable. 4. Drive new error_q. 5. Poll Done",
     "Two different error_q values (FR 3.4, AC-E2)",
     "SCB-7, SCB-4, SCB-5, SCB-6", "cond, line", "CORNER", "P2", "AC-E2"),
    ("", "Back-to-back error_q changes", "led_back_to_back_error_test",
     "1. Drive error_q=10, hold 1002. 2. Drive error_q=99, hold 1002. 3. Poll Done after each",
     "Hold >=1002 each; sequence restarts hold per error_q change (C-4, C-5)",
     "SCB-4", "fsm", "CORNER", "P2", "AC-E1, C-5"),
    ("", "Reset recovery mid-display", "led_reset_during_display_test",
     "1. Start error_q=42 display. 2. Assert reset at cycle 500. 3. Check reset values. 4. Re-run display",
     "Reset during hold (FR 3.3, AC-R1, AC-R2)",
     "assert_sel_out_reset_value, assert_seg_out_reset_value; SCB-8", "line, fsm", "CORNER", "P2", "AC-R1, AC-R2"),
    ("", "Done clears after reset", "led_done_clear_after_reset_test",
     "1. Complete display Done=1. 2. Reset. 3. Read Done expect 0",
     "AC-D2",
     "SCB-2", "line", "CORNER", "P2", "AC-D2"),
    ("", "Hold below minimum (negative)", "led_hold_below_min_negative_test",
     "1. Hold error_q for 1000 cycles (< C-4 min). 2. Drive error_q",
     "Hold 1000 cycles violates C-4; sequence drives short hold",
     "SCB-8; test expect failure or DUT not Done", "cond", "CORNER", "P2", "C-4"),
    ("", "All decimal digits 0-9 on display", "led_all_digits_0_to_9_test",
     "For d in 0..9: drive error_q with ones digit d; poll Done; compare encoding",
     "10 iterations (AC-C1, AC-B1)",
     "SCB-5; cover_seg_out_decimal_digit; COV cg_digits", "line, toggle", "CORNER", "P1", "AC-C1, AC-B1"),

    ("Integration", "End-to-end smoke", "smoke_test",
     "1. Reset. 2. Default enable. 3. error_q=42. 4. Poll Done. 5. Scoreboard pass",
     "error_q=42",
     f"SCB-1..9; {ALL_SVA_PROPERTIES}; COV all", "all", "BASIC", "P0", "AC-R1..AC-C3"),
    ("", "Full display flow with scratch pad", "full_display_flow_test",
     "1. Write enable. 2. Drive random error_q. 3. Poll Done. 4. Read scratch. 5. Read Done",
     "error_q random in [0:999_999]",
     "SCB-1, SCB-2, SCB-3, SCB-4, SCB-5, SCB-6, SCB-8", "line, branch", "BASIC", "P1", "AC-D1, AC-A1"),
    ("", "Scratch pad then LED display", "scratch_then_display_test",
     "1. Scratch write/read. 2. LED display. 3. Read Done",
     "Independent wdata and error_q",
     "SCB-3, SCB-4, SCB-5, SCB-6", "line", "BASIC", "P1", "AC-A1, AC-D1"),

    ("Corner cases", "Random regression", "random_regression_test",
     "Random APB + LED virtual sequence; error_q 80% in-range, 20% overflow; poll Done",
     "Hold >=1002 cycles (C-4); seed via +ntb_random_seed",
     f"SCB-1..9; {ALL_SVA_PROPERTIES}; COV all", "all", "CORNER", "P2", "AC-C1, AC-C2"),
    ("", "Virtual sequence stress (20x)", "virtual_seq_stress_test",
     "Run led_mux_virtual_seq 20 times with random ordering",
     "Seed-controlled (+ntb_random_seed)",
     "SCB-1..9; COV all", "all", "CORNER", "P2", "AC-C2"),
    ("", "Disable then overflow then enable", "enable_off_overflow_test",
     "1. Disable. 2. Drive error_q=1_000_050. 3. Enable. 4. Drive error_q=50. 5. Poll Done",
     "Overflow then in-range (FR 3.6, AC-E2, AC-B4)",
     "SCB-7, SCB-9", "branch, cond", "CORNER", "P2", "AC-E2, AC-B4"),
    ("", "Tight Done poll stress", "poll_until_done_stress_test",
     "Poll 0x4004 every 5 cycles while LED processes",
     "No false scoreboard failures (SCB-2, SCB-8)",
     "SCB-2, SCB-8", "fsm", "CORNER", "P2", "AC-D1"),
]

AIMS = {
    "led_reset_values_test": "Verify LED interface reset output values per spec",
    "apb_reset_defaults_test": "Verify APB register defaults after reset",
    "apb_led_enable_write_read_test": "Verify LED_enable register write/read",
    "apb_scratchpad_wr_rd_test": "Verify scratch pad write-read integrity",
    "apb_done_read_only_test": "Verify Done register is read-only",
    "apb_pready_no_wait_test": "Verify APB no-wait-state completion",
    "apb_default_enable_led_path_test": "Verify default LED_enable allows display",
    "apb_invalid_addr_test": "Verify pslerr on invalid APB address",
    "apb_done_poll_timeout_test": "Verify Done stays 0 without error_q stimulus",
    "apb_scratchpad_all_ones_test": "Verify scratch pad stores all-ones pattern",
    "apb_scratchpad_walking_one_test": "Verify scratch pad bit-walk patterns",
    "apb_enable_toggle_stress_test": "Stress LED_enable toggling with displays",
    "apb_read_during_processing_test": "Verify Done read timing during LED processing",
    "led_single_digit_zero_test": "Verify display of error_q=0",
    "led_single_digit_one_test": "Verify display of error_q=1",
    "led_decimal_42_test": "Verify BCD conversion and segment encoding for 42",
    "led_max_displayable_test": "Verify max 6-digit display value 999999",
    "led_sel_onehot_scan_test": "Verify sel_out one-hot multiplexing",
    "led_seg_active_low_test": "Verify active-low segment encoding",
    "led_hold_time_min_test": "Verify minimum 1002-cycle hold requirement",
    "led_latency_window_test": "Verify 60-80 cycle output latency window",
    "led_overflow_modulo_test": "Verify overflow modulo display",
    "led_overflow_max_test": "Verify max error_q overflow handling",
    "led_overflow_boundary_test": "Verify overflow boundary transitions",
    "led_disable_blocks_update_test": "Verify LED_enable=0 blocks seg_out update",
    "led_reenable_after_disable_test": "Verify display resumes after re-enable",
    "led_back_to_back_error_test": "Verify consecutive error_q updates",
    "led_reset_during_display_test": "Verify reset recovery mid-display",
    "led_done_clear_after_reset_test": "Verify Done resets to 0 after reset",
    "led_hold_below_min_negative_test": "Negative test for hold time below C-4",
    "led_all_digits_0_to_9_test": "Exercise all decimal digits on display",
    "smoke_test": "End-to-end sanity check of full UVM environment",
    "full_display_flow_test": "Full APB + LED flow with scratch pad",
    "scratch_then_display_test": "Scratch pad access then LED display",
    "random_regression_test": "Randomized multi-scenario regression",
    "virtual_seq_stress_test": "Stress virtual sequence coordination",
    "enable_off_overflow_test": "Overflow behavior with enable gating",
    "poll_until_done_stress_test": "Stress Done polling during processing",
}


def resolve_owner(cli_owner: str | None) -> str:
    if cli_owner and cli_owner.strip():
        return cli_owner.strip()
    owner = input("Enter testplan owner name ($owner / Test Owner): ").strip()
    if not owner:
        print("Error: owner name is required.", file=sys.stderr)
        sys.exit(1)
    return owner


def main(owner: str, tier: str = "p0"):
    rows = filter_rows_by_tier(ROWS, tier)
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb["TestPlan"]

    for row in range(ws.max_row, 1, -1):
        if row > 1:
            ws.delete_rows(row)

    headers = [
        "hvp plan", "Feature", "Sub Feature", "$owner", "$description",
        "Assertions/Cover property", "Covergroups", "Code Coverage", "Tests", "Priority",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)

    ws.cell(2, 1, PLAN_NAME)
    wrap = Alignment(wrap_text=True, vertical="top")

    for i, row in enumerate(rows, start=3):
        feat, sub, test, flow, constraints, checkers, code_cov, typ, pri, goal = row
        sva, cov, scb = parse_checkers(checkers)
        assertions = ", ".join(sva)
        covergroups = ", ".join(cov)
        desc = fmt_desc(flow, constraints, checkers)
        validate_description(desc, test)

        rank = TEST_RANK.get(test, "P2")
        values = ["", feat, sub, owner, desc, assertions, covergroups, code_cov, test, rank]
        for c, val in enumerate(values, 1):
            cell = ws.cell(i, c, val)
            cell.alignment = wrap

    widths = [22, 24, 30, 12, 60, 28, 22, 16, 34, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws2 = wb["Sheet1"]
    start = 2
    for r in range(1, 20):
        val = ws2.cell(r, 1).value
        if val and str(val).strip() == "Test name":
            start = r + 1
            break
    if ws2.max_row >= start:
        ws2.delete_rows(start, ws2.max_row - start + 1)

    checklist_headers = [
        "Test name", "Aim", "Validation level", "Test Owner", "Description",
        "Test Environment Requirement", "Test Steps", "Pass/Fail Criteria",
        "Type", "Weight", "Goal",
    ]
    for col, header in enumerate(checklist_headers, 1):
        ws2.cell(start - 1, col, header)

    for j, row in enumerate(rows, start=start):
        feat, sub, test, flow, constraints, checkers, code_cov, typ, _pri, goal = row
        rank = TEST_RANK.get(test, "P2")
        values = [
            test,
            AIMS.get(test, sub),
            "Sub-system",
            owner,
            sub,
            "UVM 1.2; led_env; APB + LED agents active (ARCHITECTURE.md)",
            flow,
            "Scoreboard pass; SVA clean; expected register/output values; coverage bin hit",
            typ,
            rank,
            goal,
        ]
        for c, val in enumerate(values, 1):
            ws2.cell(j, c, val).alignment = wrap

    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["G"].width = 50
    ws2.column_dimensions["H"].width = 40

    wb.save(OUT)
    print(f"Generated {OUT}")
    print(f"Tier: {tier} — TestPlan rows: {len(rows)}")
    print(f"Sheet1 rows: {len(rows)}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate LED_MUX_CONTROLLER_testplan.xlsx")
    parser.add_argument(
        "--owner",
        help="Verification owner for $owner and Test Owner columns (prompted if omitted)",
    )
    parser.add_argument(
        "--tier",
        choices=["p0", "p1", "p2", "all"],
        default="all",
        help="Tests to include: p0=essential (11), p1=+good-to-have (20), all=38 with Priority column (default: all)",
    )
    args = parser.parse_args()
    main(resolve_owner(args.owner), args.tier)
